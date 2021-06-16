##from win32com.client import Dispatch
##
##xlApp = Dispatch('Excel.Application')
##xlApp.Visible = True
##
##xlBook = xlApp.Workbooks.Open(r'C:\Users\021\Desktop\123\123.xlsx')
##
##for row in range(200000):
##        for col in range(0,200000):
##            xlBook.write(row + 1, col, col)
##   # file.save('123.xlsx')
##
###xlBook.Save()
##
###xlBook.Close()
##
###xlApp.Quit()

##import xlwt
##book=xlwt.Workbook(encoding='utf-8',style_compression=0)
##sheet=book.add_sheet('date',cell_overwrite_ok=True)
##for i in range(65536):
##        sheet.write(i,0,'1234')
##
##
##book.save(r'C:\Users\021\Desktop\123\7.xls')
##

import openpyxl

file = openpyxl.Workbook()  #创建一个工作簿,W大写

print(file.sheetnames)     #新建文档默认存在一个Sheet

print(file.active)

sheet = file["Sheet"]

for i in range(1,100000):
    for k in range(1,256):
        sheet.cell(row=i, column=k, value=i)
    



file.save("3.xlsx")


